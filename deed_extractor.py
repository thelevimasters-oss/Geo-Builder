"""Utility helpers for cleaning deed text before parsing bearings and distances."""

from __future__ import annotations

import argparse
import importlib
import importlib.util
import json
import logging
import os
import re
from collections import Counter
from pathlib import Path
from typing import Iterable, Iterator, List, Optional, Sequence, TextIO, Tuple

import pandas as pd

from openpyxl.utils import get_column_letter

from fractions import Fraction

import sys

_CHAR_NORMALIZE_MAP = {
    "′": "'",
    "’": "'",
    "`": "'",
    "‛": "'",
    "＇": "'",
    "¨": "'",
    "˝": '"',
    "″": '"',
    "“": '"',
    "”": '"',
    "„": '"',
    "º": "°",
    "˚": "°",
    "⁰": "°",
    "°": "°",
    "‐": "-",
    "‑": "-",
    "‒": "-",
    "–": "-",
    "—": "-",
    "―": "-",
}

_DEGREE_WORD_PATTERN = re.compile(r"(?i)\bDEG(?:REE|REES)?\b")
_MINUTE_WORD_PATTERN = re.compile(r"(?i)\bMIN(?:UTE|UTES)?\b")
_SECOND_WORD_PATTERN = re.compile(r"(?i)\bSEC(?:OND|ONDS)?\b")
_CARDINAL_PATTERN = re.compile(
    r"(?i)\b(NORTH|SOUTH|EAST|WEST)(?:\s+|-)?(EAST|WEST)?(?:ERLY)?\b"
)
_CARDINAL_ABBREV_PATTERN = re.compile(r"(?i)\b([NSEW])\.(?=\s|$)")
_LETTER_O_DEGREE_PATTERN = re.compile(r"(?<=\d)\s*[oO](?=\s*\d)")
_MULTI_SPACE_PATTERN = re.compile(r"\s+")
_CUE_WORD_PATTERN = re.compile(
    r"(?i)\b(THENCE|THEN|BEGINNING|BEGIN|ENDING|CONTAINING|WITH|ALONG|RUNNING)\b"
)
_PUNCT_GAP_PATTERN = re.compile(r"[;,]*(?=\s)")
_TRAILING_DIR_PUNCT_PATTERN = re.compile(r"(?i)\b([NSEW]{1,2})[\.,;:]+(?=\s)")
_UNIT_PUNCT_PATTERN = re.compile(
    r"(?i)\b(FEET|FT|FOOT|METERS|M|CHAINS|CHS|CHAIN|RODS|RDS|ROD)[\.,;:]+(?=\s)"
)
_HEADER_FOOTER_PATTERNS: Iterable[re.Pattern[str]] = (
    re.compile(r"(?i)^\s*page\s+\d+(?:\s+of\s+\d+)?\s*$"),
    re.compile(r"(?i)^\s*continued\s*$"),
    re.compile(r"(?i)^\s*-{2,}\s*$"),
)

_DEFAULT_MODEL_DIRNAME = "deed_ner_model"
_MODEL_META_FILENAME = "meta.json"

_ENTITY_RULER_PATTERNS = [
    {
        "label": "DEED_CALL",
        "pattern": [
            {"TEXT": {"REGEX": "(?i)^thence$"}, "OP": "?"},
            {"TEXT": {"REGEX": "(?i)^[ns](?:[ew])?$"}},
            {"TEXT": {"REGEX": r"^(?:\\d+(?:\\.\\d+)?|°|º|o|'|\"|-)$"}, "OP": "*"},
            {"TEXT": {"REGEX": "(?i)^[ns](?:[ew])?$"}, "OP": "?"},
            {"TEXT": {"REGEX": r"^\\d+(?:\\.\\d+)?$"}},
            {
                "LOWER": {
                    "IN": [
                        "feet",
                        "foot",
                        "ft",
                        "meter",
                        "meters",
                        "m",
                        "chain",
                        "chains",
                        "rod",
                        "rods",
                        "yard",
                        "yards",
                        "vara",
                        "varas",
                        "link",
                        "links",
                    ]
                },
                "OP": "?",
            },
        ],
    },
    {
        "label": "DEED_CALL",
        "pattern": [
            {"TEXT": {"REGEX": "(?i)^thence$"}, "OP": "?"},
            {
                "LOWER": {
                    "IN": [
                        "north",
                        "south",
                        "east",
                        "west",
                        "northeast",
                        "northwest",
                        "southeast",
                        "southwest",
                    ]
                }
            },
            {"TEXT": {"REGEX": r"^(?:along|with|of)$"}, "OP": "*"},
            {"TEXT": {"REGEX": r"^\\d+(?:\\.\\d+)?$"}},
            {
                "LOWER": {
                    "IN": [
                        "feet",
                        "foot",
                        "ft",
                        "meter",
                        "meters",
                        "m",
                        "chain",
                        "chains",
                        "rod",
                        "rods",
                        "yard",
                        "yards",
                        "vara",
                        "varas",
                        "link",
                        "links",
                    ]
                },
                "OP": "?",
            },
        ],
    },
]

_REGEX_FALLBACK_PATTERN = re.compile(
    r"""
    (?ix)
    \b
    (?:THENCE|THEN)?\s*
    (?:
        [NS](?:[EW])?
        (?:
            \s*\d{1,3}
            (?:\s*[°ºo]\s*\d{1,2}
                (?:\s*['′](?:\s*\d{1,2})?(?:\s*(?:"|″|''))? )?
            )?
        )?
        \s*[EW]?
        |
        (?:NORTH|SOUTH|EAST|WEST)(?:EAST|WEST)?
    )
    [^A-Z0-9]{0,10}
    \d+(?:\.\d+)?
    \s*
    (?:FEET|FOOT|FT|METERS?|M|CHAINS?|LINKS?|RODS?|YARDS?|VARAS?)
    """,
    re.VERBOSE,
)

_SPACY_AVAILABLE = importlib.util.find_spec("spacy") is not None
spacy = importlib.import_module("spacy") if _SPACY_AVAILABLE else None
_NLP_CACHE = None
_ENTITY_RULER_NAME = "deed_call_ruler"


def update_deed_model_cache(nlp=None) -> None:
    """Replace or clear the cached spaCy pipeline used for deed extraction.

    Args:
        nlp: A spaCy ``Language`` pipeline to cache. Passing ``None`` clears
            the cache so the next extraction attempt reloads from disk.
    """

    global _NLP_CACHE
    _NLP_CACHE = nlp


class NoCallsFoundError(RuntimeError):
    """Raised when no deed calls are detected after hybrid extraction."""

    def __init__(
        self,
        message: str,
        *,
        total_chars: int,
        window_count: int,
        preview: str,
        cue_words: Sequence[str],
        log_path: Optional[Path] = None,
    ) -> None:
        super().__init__(message)
        self.total_chars = total_chars
        self.window_count = window_count
        self.preview = preview
        self.cue_words = list(cue_words)
        self.log_path = Path(log_path) if log_path is not None else None


def _strip_headers_and_footers(text: str) -> str:
    lines = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if any(pattern.match(line) for pattern in _HEADER_FOOTER_PATTERNS):
            continue
        lines.append(line)
    return " ".join(lines)


def _normalize_special_chars(text: str) -> str:
    if not text:
        return ""
    chars = []
    for ch in text:
        chars.append(_CHAR_NORMALIZE_MAP.get(ch, ch))
    normalized = "".join(chars)
    normalized = _DEGREE_WORD_PATTERN.sub("°", normalized)
    normalized = _LETTER_O_DEGREE_PATTERN.sub("°", normalized)
    normalized = _MINUTE_WORD_PATTERN.sub("'", normalized)
    normalized = _SECOND_WORD_PATTERN.sub('"', normalized)
    normalized = normalized.replace("º", "°")
    return normalized


def _standardize_cardinals(text: str) -> str:
    def repl(match: re.Match[str]) -> str:
        primary = match.group(1)[0].upper()
        secondary = match.group(2)
        if secondary:
            return primary + secondary[0].upper()
        return primary

    text = _CARDINAL_PATTERN.sub(repl, text)
    text = _CARDINAL_ABBREV_PATTERN.sub(lambda m: m.group(1).upper(), text)
    return text


def _uppercase_cues(text: str) -> str:
    return _CUE_WORD_PATTERN.sub(lambda m: m.group(1).upper(), text)


def clean_text(raw: str) -> str:
    """Normalize deed text into a compact, parser-friendly representation.

    The helper removes common page headers, normalizes quotation-like
    characters, and uppercases cue words such as ``THENCE`` so that the
    downstream extractors operate on predictable input.

    >>> clean_text("Thence south 45 degrees 30 minutes west 120 feet.")
    "THENCE S45 ° 30 ' W120 feet."
    >>> clean_text("THENCE north 120 feet.")
    'THENCE N120 feet.'
    """

    if not raw:
        return ""

    text = raw.replace("\r\n", "\n").replace("\r", "\n")
    text = text.replace("-\n", "")
    text = _strip_headers_and_footers(text)
    text = _normalize_special_chars(text)
    text = _standardize_cardinals(text)
    text = _uppercase_cues(text)
    text = _TRAILING_DIR_PUNCT_PATTERN.sub(lambda m: m.group(1).upper(), text)
    text = _UNIT_PUNCT_PATTERN.sub(lambda m: m.group(1).lower(), text)
    text = _PUNCT_GAP_PATTERN.sub("", text)
    text = re.sub(r"(?<=\d),(?=\d)", ",", text)  # keep decimals/commas in numbers
    text = re.sub(r"(?<=\d)[,](?=\s)", "", text)
    text = re.sub(r"(?<=\b)([NSEW])(?=\s+\d)", lambda m: m.group(1).upper(), text)
    text = _MULTI_SPACE_PATTERN.sub(" ", text)
    return text.strip()




def iter_windows(
    text: str,
    window_chars: int = 6000,
    overlap_chars: int = 600,
) -> Iterator[Tuple[str, int]]:
    """Yield overlapping text windows to avoid silently truncating long docs.

    Args:
        text: The full text to segment.
        window_chars: Maximum number of characters per window.
        overlap_chars: Number of characters of overlap between consecutive
            windows. Must be smaller than ``window_chars``.

    Yields:
        Tuples of ``(window_text, start_offset)`` where ``start_offset`` is the
        index of the window's first character in the original ``text``.
    """

    if not text:
        return

    if window_chars <= 0:
        raise ValueError("window_chars must be a positive integer")
    if overlap_chars < 0:
        raise ValueError("overlap_chars must be zero or a positive integer")
    if overlap_chars >= window_chars:
        raise ValueError("overlap_chars must be smaller than window_chars to allow progress")

    text_length = len(text)
    start = 0
    while start < text_length:
        end = min(text_length, start + window_chars)
        yield text[start:end], start
        if end >= text_length:
            break
        start = end - overlap_chars


def get_saved_deed_model_path(base_path: Optional[Path] = None) -> Path:
    """Return the expected location of the saved deed AI model."""

    root = Path(__file__).resolve().parent if base_path is None else Path(base_path)
    return root / _DEFAULT_MODEL_DIRNAME


def expand_span_to_line(text: str, start: int, end: int, *, context_chars: int = 80) -> str:
    """Expand a span to the surrounding call line.

    Args:
        text: Source text containing the span.
        start: Start index of the span.
        end: End index of the span.
        context_chars: Number of characters to expand on each side when the
            span is not bounded by newlines.

    Returns:
        A string containing the surrounding line when the text contains
        newline-delimited lines. If no newline boundaries exist, a substring
        expanded by ``context_chars`` characters on each side is returned.
    """

    if not text:
        return ""

    text_length = len(text)
    start = max(0, min(start, text_length))
    end = max(start, min(end, text_length))

    line_start = text.rfind("\n", 0, start)
    line_end = text.find("\n", end)

    if line_start != -1 or line_end != -1:
        if line_start == -1:
            line_start = 0
        else:
            line_start += 1
        if line_end == -1:
            line_end = text_length
        return text[line_start:line_end].strip()

    window_start = max(0, start - context_chars)
    window_end = min(text_length, end + context_chars)
    return text[window_start:window_end].strip()


def _read_model_meta(model_path: Path) -> dict:
    meta_path = model_path / _MODEL_META_FILENAME
    if not meta_path.exists():
        return {}
    try:
        with meta_path.open("r", encoding="utf-8") as fh:
            return json.load(fh)
    except Exception:
        return {}


def _extract_labels_from_meta(meta: dict) -> List[str]:
    labels: List[str] = []

    possible_sources = []
    labels_section = meta.get("labels")
    if isinstance(labels_section, dict):
        possible_sources.extend(
            value
            for key in ("ner", "ents")
            for value in [labels_section.get(key)]
        )

    components = meta.get("components")
    if isinstance(components, dict):
        ner_component = components.get("ner")
        if isinstance(ner_component, dict):
            possible_sources.append(ner_component.get("labels"))

    ner_section = meta.get("ner")
    if isinstance(ner_section, dict):
        possible_sources.append(ner_section.get("labels"))

    for source in possible_sources:
        if isinstance(source, list):
            for label in source:
                if isinstance(label, str):
                    labels.append(label)

    if not labels:
        return []

    # Preserve original ordering while removing duplicates.
    seen = set()
    unique_labels: List[str] = []
    for label in labels:
        if label not in seen:
            seen.add(label)
            unique_labels.append(label)
    return unique_labels


def _extract_labels_from_spacy(nlp) -> List[str]:  # pragma: no cover - spaCy optional
    labels: List[str] = []
    if nlp is None:
        return labels
    try:
        pipe_names = getattr(nlp, "pipe_names", [])
    except Exception:
        pipe_names = []
    if "ner" not in pipe_names:
        return labels
    try:
        ner = nlp.get_pipe("ner")
        ner_labels = getattr(ner, "labels", None)
    except Exception:
        ner_labels = None
    if ner_labels:
        for label in ner_labels:
            if isinstance(label, str):
                labels.append(label)
    if not labels:
        return []
    seen = set()
    unique_labels: List[str] = []
    for label in labels:
        if label not in seen:
            seen.add(label)
            unique_labels.append(label)
    return unique_labels


def load_saved_deed_model_meta(model_path: Optional[Path] = None) -> dict:
    resolved_path = get_saved_deed_model_path() if model_path is None else model_path
    if not isinstance(resolved_path, Path):
        resolved_path = Path(resolved_path)
    return _read_model_meta(resolved_path)


def check_saved_deed_model(
    model_path: Optional[Path] = None,
    *,
    stream: Optional[TextIO] = None,
) -> List[str]:
    """Print diagnostic information about the saved deed AI model."""

    output = stream or sys.stdout
    resolved_path = get_saved_deed_model_path() if model_path is None else model_path
    if not isinstance(resolved_path, Path):
        resolved_path = Path(resolved_path)
    print(f"Deed AI model path: {resolved_path.resolve()}", file=output)

    labels: List[str] = []
    nlp = None
    try:  # pragma: no cover - spaCy optional
        import spacy  # type: ignore
    except Exception as exc:  # pragma: no cover - spaCy optional
        print(f"spaCy unavailable ({exc}). Using saved metadata only.", file=output)
    else:  # pragma: no cover - spaCy optional
        try:
            nlp = spacy.load(resolved_path)
        except Exception as exc:
            print(
                f"spaCy could not load saved model ({exc}). Falling back to metadata.",
                file=output,
            )
            nlp = None
        labels = _extract_labels_from_spacy(nlp)
    if not labels:
        meta = load_saved_deed_model_meta(resolved_path)
        labels = _extract_labels_from_meta(meta)
    if not labels:
        labels = ["DEED_CALL"]

    print("Loaded saved deed AI model.", file=output)
    print(f"NER labels: {labels}", file=output)
    return labels


def _ensure_entity_ruler(nlp) -> None:  # pragma: no cover - spaCy optional
    if nlp is None:
        return
    existing = []
    try:
        existing = list(getattr(nlp, "pipe_names", []))
    except Exception:
        existing = []
    if _ENTITY_RULER_NAME in existing:
        return
    try:
        before = "ner" if "ner" in existing else None
        ruler = nlp.add_pipe("entity_ruler", name=_ENTITY_RULER_NAME, before=before)
    except Exception:
        return
    try:
        ruler.add_patterns(_ENTITY_RULER_PATTERNS)
    except Exception:
        pass


def _ensure_ner_label(nlp) -> None:  # pragma: no cover - spaCy optional
    if nlp is None:
        return
    try:
        pipe_names = getattr(nlp, "pipe_names", [])
    except Exception:
        pipe_names = []
    if "ner" not in pipe_names:
        return
    try:
        ner = nlp.get_pipe("ner")
    except Exception:
        return
    add_label = getattr(ner, "add_label", None)
    if not callable(add_label):
        return
    try:
        add_label("DEED_CALL")
    except Exception:
        pass


def _get_deed_nlp():  # pragma: no cover - spaCy optional
    global _NLP_CACHE
    if _NLP_CACHE is not None:
        return _NLP_CACHE
    if spacy is None:
        _NLP_CACHE = None
        return _NLP_CACHE
    model_path = get_saved_deed_model_path()
    nlp = None
    if model_path.exists():
        try:
            nlp = spacy.load(model_path)
        except Exception:
            nlp = None
    if nlp is None:
        try:
            nlp = spacy.blank("en")
        except Exception:
            nlp = None
    if nlp is not None:
        _ensure_entity_ruler(nlp)
        _ensure_ner_label(nlp)
    _NLP_CACHE = nlp
    return _NLP_CACHE


def _should_use_ner() -> bool:
    disable_value = os.getenv("DEED_EXTRACTOR_DISABLE_NER")
    if disable_value is None:
        return True
    disable_value = disable_value.strip().lower()
    return disable_value not in {"1", "true", "yes", "on"}


def _iter_regex_matches(window: str, *, offset: int):
    for match in _REGEX_FALLBACK_PATTERN.finditer(window):
        start = offset + match.start()
        end = offset + match.end()
        yield start, end


def _format_no_calls_message(
    *,
    total_chars: int,
    window_count: int,
    preview: str,
    cue_words: Sequence[str],
) -> str:
    cues = ", ".join(cue_words) if cue_words else "none"
    preview_text = preview or ""
    return (
        "No deed calls were found after applying NER and regex.\n"
        f"Characters/windows: {total_chars}/{window_count}\n"
        f"Cue words: {cues}\n"
        f"Cleaned preview (first 200 chars): {preview_text}"
    )


def _write_extraction_log(
    log_path: Path,
    *,
    total_chars: int,
    window_count: int,
    extractor_counts: Counter,
    rows: List[dict],
) -> None:
    """Persist extraction diagnostics to a log file."""

    log_lines = [
        f"total_chars={total_chars}",
        f"window_count={window_count}",
        f"result_count={len(rows)}",
    ]

    for source in sorted(extractor_counts):
        log_lines.append(f"{source}_count={extractor_counts[source]}")

    log_lines.append("rows:")
    for index, row in enumerate(rows, start=1):
        snippet = row.get("text", "").replace("\n", " ")
        source = row.get("source", "unknown")
        start = row.get("start", "?")
        end = row.get("end", "?")
        log_lines.append(
            f"{index}\tsource={source}\tstart={start}\tend={end}\ttext={snippet}"
        )

    try:
        log_path.parent.mkdir(parents=True, exist_ok=True)
        with log_path.open("w", encoding="utf-8") as log_file:
            log_file.write("\n".join(log_lines) + "\n")
    except Exception:
        # Logging must never break the extractor; failures are silently ignored.
        pass


def extract_calls_hybrid(
    text: str,
    *,
    window_chars: int = 6000,
    overlap_chars: int = 600,
    log_path: Optional[Path] = None,
    use_ner: Optional[bool] = None,
    regex_fallback: bool = True,
) -> List[dict]:
    """Locate deed calls using spaCy NER with an optional regex fallback.

    Args:
        text: Raw deed description text to analyze.
        window_chars: Maximum number of characters per sliding window.
        overlap_chars: Number of characters that overlap between windows.
        log_path: Destination for the extraction diagnostics log.
        use_ner: Force enabling/disabling the spaCy model. ``None`` keeps the
            environment-controlled default.
        regex_fallback: When ``True`` the regex extractor supplements missing
            NER spans. Disable to require NER results only.
    """

    cleaned = clean_text(text)
    total_chars = len(cleaned)
    cue_words_found = sorted({word.upper() for word in _CUE_WORD_PATTERN.findall(cleaned)})
    preview = cleaned[:200]
    if len(cleaned) > 200:
        preview += "..."

    resolved_log_path = Path(log_path) if log_path is not None else Path("deed_extractor.log")

    use_ner_flag = _should_use_ner() if use_ner is None else bool(use_ner)
    nlp = _get_deed_nlp() if use_ner_flag else None
    if use_ner is True and nlp is None:
        raise RuntimeError("spaCy model is unavailable for NER extraction.")

    results: List[dict] = []
    seen_spans = set()
    extractor_counts: Counter = Counter({"ner": 0, "regex": 0})
    log_rows: List[dict] = []
    window_count = 0

    if not cleaned:
        _write_extraction_log(
            resolved_log_path,
            total_chars=total_chars,
            window_count=window_count,
            extractor_counts=extractor_counts,
            rows=log_rows,
        )
        message = _format_no_calls_message(
            total_chars=total_chars,
            window_count=window_count,
            preview=preview,
            cue_words=cue_words_found,
        )
        raise NoCallsFoundError(
            message,
            total_chars=total_chars,
            window_count=window_count,
            preview=preview,
            cue_words=cue_words_found,
            log_path=resolved_log_path,
        )

    for window_text, start_offset in iter_windows(
        cleaned, window_chars=window_chars, overlap_chars=overlap_chars
    ):
        window_count += 1
        window_spans: List[Tuple[int, int, str]] = []
        if nlp is not None:
            try:
                doc = nlp(window_text)
            except Exception:
                doc = None
            if doc is not None:
                for ent in getattr(doc, "ents", []):
                    if ent.label_ != "DEED_CALL":
                        continue
                    span_start = start_offset + int(ent.start_char)
                    span_end = start_offset + int(ent.end_char)
                    if span_end <= span_start:
                        continue
                    window_spans.append((span_start, span_end, "ner"))

        if regex_fallback and not any(source == "ner" for _, _, source in window_spans):
            for span_start, span_end in _iter_regex_matches(window_text, offset=start_offset):
                window_spans.append((span_start, span_end, "regex"))

        for span_start, span_end, source in window_spans:
            if (span_start, span_end) in seen_spans:
                continue
            seen_spans.add((span_start, span_end))
            span_text = cleaned[span_start:span_end].strip()
            if not span_text:
                continue
            row = {
                "text": span_text,
                "start": span_start,
                "end": span_end,
                "label": "DEED_CALL",
                "source": source,
            }
            results.append(row)
            log_rows.append(row)
            extractor_counts[source] += 1

    results.sort(key=lambda item: item.get("start", 0))

    _write_extraction_log(
        resolved_log_path,
        total_chars=total_chars,
        window_count=window_count,
        extractor_counts=extractor_counts,
        rows=log_rows,
    )

    if not results:
        message = _format_no_calls_message(
            total_chars=total_chars,
            window_count=window_count,
            preview=preview,
            cue_words=cue_words_found,
        )
        raise NoCallsFoundError(
            message,
            total_chars=total_chars,
            window_count=window_count,
            preview=preview,
            cue_words=cue_words_found,
            log_path=resolved_log_path,
        )

    return results


_BEARING_PATTERN = re.compile(
    r"""
    ^\s*
    (?P<primary>NORTH|SOUTH|N|S)
    (?:\s+|-)*
    (?P<angle>.+?)
    (?:\s+|-)*
    (?P<secondary>EAST|WEST|E|W)
    \s*$
    """,
    re.IGNORECASE | re.VERBOSE,
)

_NUMERIC_TOKEN_PATTERN = re.compile(
    r"(?P<num>[-+]?\d+(?:\s+\d+/\d+)?|\d+/\d+|[-+]?\d*\.\d+)(?:\s*(?P<sym>°|''|\"|['′]))?",
)


def _parse_fractional_number(text: str) -> Fraction:
    text = text.strip()
    if not text:
        raise ValueError("Expected a numeric value")
    cleaned = text.replace(",", "")
    parts = cleaned.split()
    if not parts:
        raise ValueError("Expected a numeric value")
    total = Fraction(0, 1)
    for part in parts:
        try:
            total += Fraction(part)
        except (ValueError, ZeroDivisionError) as exc:  # pragma: no cover - defensive
            raise ValueError(f"Invalid numeric component: {part}") from exc
    return total


def _coerce_fraction(value: Fraction) -> float | int:
    if value.denominator == 1:
        return int(value)
    return float(value)


def parse_bearing(value: str) -> dict:
    """Parse a quadrant bearing string into its components.

    Args:
        value: Raw bearing text such as ``"S 48° 45' W"``.

    Returns:
        A dictionary containing the primary and secondary quadrant along with
        normalized degree, minute, and second values.
    >>> parse_bearing("S 48° 30' E")
    {'quadrant_1': 'S', 'quadrant_2': 'E', 'degrees': 48, 'minutes': 30, 'seconds': 0}
    >>> parse_bearing("N 45 30' 15\\\" W")
    {'quadrant_1': 'N', 'quadrant_2': 'W', 'degrees': 45, 'minutes': 30, 'seconds': 15}
    >>> parse_bearing('N 12 1/2 E')
    {'quadrant_1': 'N', 'quadrant_2': 'E', 'degrees': 12, 'minutes': 30, 'seconds': 0}
    """

    if not value:
        raise ValueError("Bearing text is required")

    normalized = _normalize_special_chars(value)
    normalized = normalized.strip()
    if not normalized:
        raise ValueError("Bearing text is required")

    match = _BEARING_PATTERN.match(normalized)
    if not match:
        raise ValueError(f"Unsupported bearing format: {value!r}")

    primary = match.group("primary")[0].upper()
    secondary = match.group("secondary")[0].upper()
    angle_text = match.group("angle") or ""

    tokens = list(_NUMERIC_TOKEN_PATTERN.finditer(angle_text))
    if not tokens:
        raise ValueError(f"Could not parse bearing angle: {value!r}")

    degrees_value: Optional[Fraction] = None
    minutes_value: Optional[Fraction] = None
    seconds_value: Optional[Fraction] = None

    for token in tokens:
        number_text = token.group("num")
        symbol = token.group("sym")
        if not number_text:
            continue
        magnitude = _parse_fractional_number(number_text)
        if symbol == "°":
            degrees_value = magnitude
        elif symbol in {"'", "′"}:
            minutes_value = magnitude
        elif symbol in {"\"", "''"}:
            seconds_value = magnitude
        else:
            if degrees_value is None:
                degrees_value = magnitude
            elif minutes_value is None:
                minutes_value = magnitude
            elif seconds_value is None:
                seconds_value = magnitude

    if degrees_value is None:
        raise ValueError(f"Bearing is missing degrees: {value!r}")

    if minutes_value is None:
        minutes_value = Fraction(0, 1)
    if seconds_value is None:
        seconds_value = Fraction(0, 1)

    total_seconds = (
        degrees_value * Fraction(3600, 1)
        + minutes_value * Fraction(60, 1)
        + seconds_value
    )

    if total_seconds < 0:
        raise ValueError("Quadrant bearings must be non-negative")

    degrees = total_seconds // 3600
    remainder = total_seconds - degrees * 3600
    minutes = remainder // 60
    seconds = remainder - minutes * 60

    return {
        "quadrant_1": primary,
        "quadrant_2": secondary,
        "degrees": _coerce_fraction(Fraction(degrees, 1)),
        "minutes": _coerce_fraction(Fraction(minutes, 1)),
        "seconds": _coerce_fraction(seconds),
    }


_DISTANCE_UNIT_TO_FEET = {
    "foot": Fraction(1, 1),
    "feet": Fraction(1, 1),
    "ft": Fraction(1, 1),
    "rod": Fraction(33, 2),
    "rods": Fraction(33, 2),
    "rd": Fraction(33, 2),
    "rds": Fraction(33, 2),
    "pole": Fraction(33, 2),
    "poles": Fraction(33, 2),
    "perch": Fraction(33, 2),
    "perches": Fraction(33, 2),
    "chain": Fraction(66, 1),
    "chains": Fraction(66, 1),
    "ch": Fraction(66, 1),
    "chs": Fraction(66, 1),
    "vara": Fraction(25, 9),
    "varas": Fraction(25, 9),
}


def normalize_distance(value: str) -> float:
    """Convert deed distance strings into feet.

    Args:
        value: Distance text such as ``"28 1/2 rods"``.

    Returns:
        The numeric distance in feet.
    >>> normalize_distance('28 1/2 rods')
    470.25
    >>> normalize_distance('15.25 ft.')
    15.25
    >>> normalize_distance('3 chains')
    198.0
    """

    if not value:
        raise ValueError("Distance text is required")

    normalized = _normalize_special_chars(value).strip().lower()
    if not normalized:
        raise ValueError("Distance text is required")

    number_match = re.search(
        r"[-+]?(?:\d*\.\d+|\d+/\d+|\d+(?:\s+\d+/\d+)?)",
        normalized,
    )
    if not number_match:
        raise ValueError(f"Could not find numeric distance in {value!r}")

    magnitude_text = number_match.group(0)
    magnitude = _parse_fractional_number(magnitude_text)

    unit_text = normalized[number_match.end() :].strip().strip(".;,)")
    unit_match = re.match(r"([a-z]+)", unit_text)
    unit_key = unit_match.group(1) if unit_match else "ft"

    conversion = _DISTANCE_UNIT_TO_FEET.get(unit_key)
    if conversion is None:
        raise ValueError(f"Unsupported distance unit: {unit_key!r}")

    feet = magnitude * conversion
    return float(feet)


def validate_training_corpus(
    path: Path | str,
    *,
    lang: Optional[str] = None,
    stream: Optional[TextIO] = None,
) -> dict:
    """Validate a spaCy DocBin training corpus.

    Args:
        path: Location of the serialized DocBin training data.
        lang: Optional spaCy language code used to initialize the vocab.
        stream: Destination for diagnostic output. Defaults to ``sys.stdout``.

    Returns:
        A dictionary containing summary statistics about the corpus.

    Raises:
        RuntimeError: If spaCy is not available in the runtime environment.
        FileNotFoundError: If ``path`` does not exist.
        ValueError: If the corpus contains no entities or the DEED_CALL label
            is missing.
    """

    if not _SPACY_AVAILABLE:  # pragma: no cover - spaCy optional
        raise RuntimeError("spaCy is required to validate training corpora.")

    if path is None:
        raise ValueError("Training corpus path is required")

    resolved_path = Path(path)
    if not resolved_path.exists():
        raise FileNotFoundError(f"Training corpus not found: {resolved_path}")

    from spacy.tokens import DocBin  # pragma: no cover - spaCy optional

    output = stream or sys.stdout

    doc_bin = DocBin().from_disk(resolved_path)

    detected_lang: Optional[str] = None
    try:  # pragma: no cover - spaCy optional
        detected_lang = doc_bin.attrs.get("lang")  # type: ignore[attr-defined]
    except Exception:
        detected_lang = None

    vocab_lang = lang or detected_lang or "en"

    try:  # pragma: no cover - spaCy optional
        vocab_source = spacy.blank(vocab_lang)
    except Exception:
        vocab_source = spacy.blank("en")

    docs = list(doc_bin.get_docs(vocab_source.vocab))

    label_counts: Counter[str] = Counter()
    total_spans = 0
    total_span_chars = 0
    dropped_examples: List[Tuple[int, str, str]] = []

    for doc_index, doc in enumerate(docs):
        for ent in getattr(doc, "ents", ()):  # pragma: no branch - defensive
            label_counts[ent.label_] += 1
            span_length = int(ent.end_char) - int(ent.start_char)
            total_span_chars += span_length
            total_spans += 1
            contracted = doc.char_span(
                int(ent.start_char),
                int(ent.end_char),
                label=ent.label_,
                alignment_mode="contract",
            )
            if contracted is None:
                preview = doc.text[int(ent.start_char) : int(ent.end_char)].strip()
                dropped_examples.append((doc_index, ent.label_, preview))

    if total_spans == 0:
        raise ValueError("Training corpus does not contain any entity spans.")

    deed_call_count = label_counts.get("DEED_CALL", 0)
    if deed_call_count == 0:
        raise ValueError("Training corpus is missing DEED_CALL annotations.")

    if dropped_examples:
        sample_doc, sample_label, sample_text = dropped_examples[0]
        raise ValueError(
            "Encountered spans that cannot be aligned with alignment_mode='contract'. "
            f"Example: doc {sample_doc}, label {sample_label!r}, text {sample_text!r}"
        )

    average_span_length = total_span_chars / total_spans if total_spans else 0.0

    print(f"Validating training corpus: {resolved_path}", file=output)
    print(f"Documents: {len(docs)}", file=output)
    print(f"Total entities: {total_spans}", file=output)
    print(f"Average span length (chars): {average_span_length:.2f}", file=output)
    print("Label frequencies:", file=output)
    for label, count in sorted(label_counts.items(), key=lambda item: (-item[1], item[0])):
        print(f"  {label}: {count}", file=output)
    print("All entity spans align with alignment_mode='contract'.", file=output)

    return {
        "documents": len(docs),
        "total_entities": total_spans,
        "average_span_length": average_span_length,
        "label_counts": dict(label_counts),
    }


_VALID_MODELS = {"hybrid", "ner", "regex"}


def extract_calls_with_model(
    model: str,
    text: str,
    *,
    log_path: Optional[Path] = None,
) -> List[dict]:
    """Dispatch deed extraction according to the selected model."""

    normalized = (model or "").strip().lower()
    if not normalized:
        normalized = "hybrid"
    if normalized == "hybrid":
        return extract_calls_hybrid(text, log_path=log_path)
    if normalized == "regex":
        return extract_calls_hybrid(text, log_path=log_path, use_ner=False)
    if normalized == "ner":
        return extract_calls_hybrid(
            text,
            log_path=log_path,
            use_ner=True,
            regex_fallback=False,
        )
    raise ValueError(f"Unsupported model: {model}")


def _load_pdf_text(pdf_path: Path) -> str:
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")
    if pdf_path.is_dir():
        raise IsADirectoryError(f"Expected a PDF file, got directory: {pdf_path}")

    logger = logging.getLogger(__name__)
    text_chunks: List[str] = []

    try:  # pragma: no cover - optional dependency
        import pdfplumber  # type: ignore
    except Exception:
        pdfplumber = None

    if pdfplumber is not None:  # pragma: no cover - optional dependency
        try:
            with pdfplumber.open(str(pdf_path)) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text() or ""
                    if page_text:
                        text_chunks.append(page_text)
        except Exception as exc:
            logger.debug("pdfplumber extraction failed: %s", exc)
        else:
            combined = "\n".join(text_chunks).strip()
            if combined:
                return combined
            text_chunks.clear()

    try:  # pragma: no cover - optional dependency
        import fitz  # type: ignore
    except Exception:
        fitz = None

    if fitz is not None:  # pragma: no cover - optional dependency
        try:
            document = fitz.open(str(pdf_path))
        except Exception as exc:
            logger.debug("PyMuPDF open failed for %s: %s", pdf_path, exc)
        else:
            try:
                for page in document:
                    try:
                        page_text = page.get_text()
                    except Exception as exc:
                        logger.debug("PyMuPDF get_text failed: %s", exc)
                        continue
                    if page_text:
                        text_chunks.append(page_text)
            finally:
                document.close()
            combined = "\n".join(text_chunks).strip()
            if combined:
                return combined

    raise RuntimeError(
        "Unable to extract text from PDF. Install pdfplumber or PyMuPDF (pymupdf)."
    )


def _load_cli_text(args: argparse.Namespace) -> str:
    if getattr(args, "text", None) is not None:
        return str(args.text)
    pdf_value = getattr(args, "pdf", None)
    if pdf_value is not None:
        return _load_pdf_text(Path(pdf_value))
    raise ValueError("Either --text or --pdf must be provided")


def _build_cli_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Extract deed calls from text or PDF sources.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--model",
        type=str,
        default="hybrid",
        help="Extraction model to use (hybrid, ner, regex).",
    )
    source_group = parser.add_mutually_exclusive_group(required=True)
    source_group.add_argument(
        "--pdf",
        type=Path,
        help="Path to a deed PDF file to analyze.",
    )
    source_group.add_argument(
        "--text",
        type=str,
        help="Raw deed text to analyze.",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Optional Excel workbook path for extracted calls.",
    )
    parser.add_argument(
        "--template",
        type=Path,
        default=None,
        help="Excel template used to order columns when writing output.",
    )
    parser.add_argument(
        "--log",
        type=Path,
        default=Path("deed_extractor.log"),
        help="Destination for the extraction diagnostics log.",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Enable verbose logging for troubleshooting.",
    )
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = _build_cli_parser()
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.debug else logging.INFO,
        format="%(message)s",
    )
    logger = logging.getLogger(__name__)

    model_choice = (args.model or "").strip().lower()
    if not model_choice:
        model_choice = "hybrid"
    if model_choice not in _VALID_MODELS:
        print(
            f"Unsupported model {args.model!r}. Choose from: {', '.join(sorted(_VALID_MODELS))}.",
            file=sys.stderr,
        )
        return 4

    try:
        text_value = _load_cli_text(args)
    except (FileNotFoundError, IsADirectoryError) as exc:
        print(str(exc), file=sys.stderr)
        return 3
    except RuntimeError as exc:
        print(str(exc), file=sys.stderr)
        return 3

    log_path = Path(args.log) if args.log is not None else Path("deed_extractor.log")

    try:
        calls = extract_calls_with_model(model_choice, text_value, log_path=log_path)
    except NoCallsFoundError as exc:
        message_lines = [str(exc)]
        if exc.log_path is not None:
            message_lines.append(f"Extraction log written to: {exc.log_path}")
        print("\n".join(message_lines), file=sys.stderr)
        return 2
    except RuntimeError as exc:
        print(str(exc), file=sys.stderr)
        return 3
    except Exception as exc:  # pragma: no cover - defensive
        logger.debug("Unexpected extraction failure", exc_info=True)
        print(f"Extraction failed: {exc}", file=sys.stderr)
        return 1

    if not calls:
        print("No deed calls were found.", file=sys.stderr)
        return 2

    df = pd.DataFrame(calls)
    if not df.empty:
        df = df.rename(
            columns={
                "text": "RawCall",
                "start": "CharStart",
                "end": "CharEnd",
                "label": "Label",
                "source": "Extractor",
            }
        )

    if args.out is not None:
        try:
            if args.template is not None:
                write_calls_xlsx(df, args.template, args.out)
            else:
                df.to_excel(args.out, index=False)
        except FileNotFoundError as exc:
            print(str(exc), file=sys.stderr)
            return 3
        except OSError as exc:
            print(f"Failed to write Excel file {args.out}: {exc}", file=sys.stderr)
            return 3
        except ValueError as exc:
            print(f"Failed to prepare Excel output: {exc}", file=sys.stderr)
            return 3
        print(f"Wrote {len(df)} call(s) to {args.out}")
    else:
        print(json.dumps(df.to_dict(orient="records"), indent=2))

    print(f"Extraction log written to: {log_path}")
    return 0


def write_calls_xlsx(df: "pd.DataFrame", template_path: os.PathLike[str] | str, out_path: os.PathLike[str] | str) -> None:
    """Write a deed calls workbook that follows a template layout."""

    with pd.ExcelFile(template_path) as template_workbook:
        sheet_name = template_workbook.sheet_names[0] if template_workbook.sheet_names else 0
        template_columns = list(
            pd.read_excel(template_workbook, sheet_name=sheet_name, nrows=0).columns
        )

    if isinstance(sheet_name, int):
        sheet_name = "Sheet1"

    if not template_columns:
        raise ValueError("Template workbook must contain at least one column header")

    ordered = df.reindex(columns=template_columns)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        ordered.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]

        worksheet.freeze_panes = "A2"

        last_column_index = len(template_columns) if template_columns else max(len(ordered.columns), 1)
        last_row_index = len(ordered.index) + 1 if len(ordered.index) else 1
        last_column_letter = get_column_letter(last_column_index)
        worksheet.auto_filter.ref = f"A1:{last_column_letter}{last_row_index}"

        if "Distance_FT" in ordered.columns:
            distance_col_index = ordered.columns.get_loc("Distance_FT") + 1
            if last_row_index > 1:
                for cell in worksheet.iter_rows(
                    min_row=2,
                    max_row=last_row_index,
                    min_col=distance_col_index,
                    max_col=distance_col_index,
                ):
                    cell[0].number_format = "#,##0.00"


def canonicalize_df(df: "pd.DataFrame") -> "pd.DataFrame":
    """Return a DataFrame that matches the canonical deed call schema.

    The canonical schema enforces a predictable column order, assigns
    sequential identifiers, and annotates rows that correspond to the
    beginning or ending of a deed description.
    """

    required_columns = [
        "DocID",
        "Sequence",
        "RawCall",
        "Bearing",
        "AngleDeg",
        "AngleMin",
        "AngleSec",
        "Quadrant1",
        "Quadrant2",
        "Distance_FT",
        "Unit_Original",
        "Monument",
        "Start_End",
        "SourcePage",
        "CharStart",
        "CharEnd",
        "Extractor",
        "Confidence",
    ]

    canonical = df.copy()

    # Ensure columns needed for sorting exist before sequencing.
    if "SourcePage" not in canonical.columns:
        canonical["SourcePage"] = None
    if "CharStart" not in canonical.columns:
        canonical["CharStart"] = None

    if not canonical.empty:
        canonical = canonical.sort_values(
            by=["SourcePage", "CharStart"],
            kind="mergesort",
            na_position="last",
        ).reset_index(drop=True)
        canonical["Sequence"] = range(1, len(canonical) + 1)
    else:
        canonical["Sequence"] = pd.Series(dtype="int64")

    # Determine which rows correspond to the beginning or end of the calls.
    start_end_markers: List[str] = [""] * len(canonical)
    if "RawCall" in canonical.columns and not canonical.empty:
        raw_call = canonical["RawCall"].fillna("").astype(str)

        begin_mask = raw_call.str.contains(r"\bBEGINNING\b", case=False, na=False)
        if begin_mask.any():
            begin_index = begin_mask[begin_mask].index[0]
            start_end_markers[begin_index] = "BEGIN"

        end_mask = raw_call.str.contains(r"to\s+the\s+beginning", case=False, na=False)
        if end_mask.any():
            end_index = end_mask[end_mask].index[-1]
            if start_end_markers[end_index]:
                start_end_markers[end_index] = f"{start_end_markers[end_index]};END"
            else:
                start_end_markers[end_index] = "END"

    canonical["Start_End"] = start_end_markers

    for column in required_columns:
        if column not in canonical.columns:
            canonical[column] = None

    canonical = canonical[required_columns]
    return canonical


__all__ = [
    "clean_text",
    "iter_windows",
    "get_saved_deed_model_path",
    "load_saved_deed_model_meta",
    "check_saved_deed_model",
    "extract_calls_hybrid",
    "extract_calls_with_model",
    "parse_bearing",
    "normalize_distance",
    "write_calls_xlsx",
    "canonicalize_df",
    "validate_training_corpus",
    "main",
    "NoCallsFoundError",
]


if __name__ == "__main__":  # pragma: no cover - CLI entry point
    raise SystemExit(main())

